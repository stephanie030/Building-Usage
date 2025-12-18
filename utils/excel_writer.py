"""
Excel 輸出工具
"""
import os
from datetime import datetime
from typing import Dict, List
import openpyxl


# Excel 表格欄位寬度對應表
CELL_LENGTH_MAPPING = {
    "_id": 10,
    "發照日期": 15,
    "執照類別": 10,
    "使照掛號日期": 15,
    "開工日期": 15,
    "竣工日期": 15,
    "竣工期限": 20,
    "竣工展期至": 20,
    "申報進度": 20,
    "核發執照字號": 35,
    "原領執照字號": 35,
    "變更設計次數": 15,
    "基地面積": 15,
    "建築面積": 15,
    "總樓地板面積": 15,
    "設計建蔽率": 15,
    "設計容積率": 15,
    "建築物高度": 15,
    "地下避難面積": 15,
    "法定空地面積": 15,
    "建造類別": 10,
    "構造別": 18,
    "棟數": 10,
    "幢數": 10,
    "地上層數": 10,
    "地下層數": 10,
    "戶數": 10,
    "起造人代表人": 20,
    "設計人": 15,
    "設計人事務所": 20,
    "監造人": 15,
    "監造人事務所": 20,
    "承造人": 20,
    "承造人營造廠": 20,
    "土地使用分區": 25,
    "建築物用途": 25,
    "工程造價": 15,
    "樓層概要": 25,
    "地號": 25,
    "門牌": 25,
    "備註資料": 25,
}


def adjust_width(ws) -> None:
    """調整 Excel 表格欄位寬度"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        header = column[0].value
        adjusted_width = CELL_LENGTH_MAPPING.get(header, max_length)
        ws.column_dimensions[column_letter].width = adjusted_width


class ExcelWriter:
    """Excel 檔案寫入器"""

    def __init__(self, city_name: str, save_dir: str):
        self.city_name = city_name
        self.save_dir = save_dir
        os.makedirs(save_dir, exist_ok=True)

        self.workbooks = {
            "建造執照": openpyxl.Workbook(),
            "使用執照": openpyxl.Workbook(),
        }
        self.worksheets = {
            "建造執照": {},
            "使用執照": {},
        }

    def write_item(self, item: Dict) -> None:
        """寫入一筆資料"""
        date_obj = item.get('date')
        data = item.get('data')

        if not data or not date_obj:
            return

        # 判斷執照類別
        data_type = "建造執照" if data.get("執照類別") == '建造執照' else "使用執照"

        workbook = self.workbooks[data_type]
        worksheets = self.worksheets[data_type]

        # 根據年月生成工作表名稱
        worksheet_name = f"{date_obj.year}-{date_obj.month:02d}"

        first_add = False
        if worksheet_name not in worksheets:
            if worksheet_name in workbook.sheetnames:
                workbook.remove(workbook[worksheet_name])
            worksheets[worksheet_name] = workbook.create_sheet(title=worksheet_name)
            first_add = True

        current_worksheet = worksheets[worksheet_name]

        # 第一次添加時加入標題行
        if first_add:
            header = list(data.keys())
            current_worksheet.append(header)

        values = list(data.values())
        current_worksheet.append(values)

    def save(self) -> Dict[str, str]:
        """保存 Excel 檔案並返回檔案路徑"""
        saved_files = {}

        for data_type in ["建造執照", "使用執照"]:
            workbook = self.workbooks[data_type]
            save_path = os.path.join(self.save_dir, f"{self.city_name}_{data_type}.xlsx")

            if len(workbook.sheetnames) > 1:
                if "Sheet" in workbook.sheetnames:
                    workbook.remove(workbook["Sheet"])
                workbook._sheets.sort(key=lambda ws: datetime.strptime(ws.title, "%Y-%m"))

            # 調整欄位寬度
            for ws in workbook.worksheets:
                adjust_width(ws)

            workbook.save(save_path)
            saved_files[data_type] = save_path

        return saved_files

    def get_stats(self) -> Dict:
        """取得統計資訊"""
        stats = {}
        for data_type in ["建造執照", "使用執照"]:
            total_rows = 0
            for ws in self.workbooks[data_type].worksheets:
                if ws.title != "Sheet":
                    total_rows += ws.max_row - 1 if ws.max_row > 1 else 0
            stats[data_type] = total_rows
        return stats
